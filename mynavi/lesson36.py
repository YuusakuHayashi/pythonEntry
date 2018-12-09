import openpyxl as excel

# wb = excel.Workbook()
# ws = wb.active
# ws["A1"] = "さようなら"
# wb.save("hello.xlsx")

wb = excel.Workbook()
ws = wb.active

for i in range(1, 10):
    for j in range(1, 10):
        v = i * j
        ws.cell(column=j, row=i, value=v)

wb.save("kuku.xlsx")
